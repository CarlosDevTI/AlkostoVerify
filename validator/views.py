from django.shortcuts import render, redirect
import random
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from .oracle_service import get_user_data
from .models import ValidationRecord

logger = logging.getLogger(__name__)

#! --- Datos Falsos para Generar Opciones ---
DUMMY_DATA = {
    'ciudad_exp': ['Bogotá D.C.', 'Medellín', 'Cali', 'Barranquilla', 'Cartagena', 'Soacha', 'Cúcuta', 'Ibagué', 'Bucaramanga', 'Pereira'],
    'tipocredito': ['Libre Inversión', 'Hipotecario', 'Vehículo', 'Educativo', 'Microcrédito', 'Tarjeta de Crédito'],
    'direccion': ['Calle 42 # 15 - 24', 'Avenida 40 # 27 -42', 'Carrera 10 # 20-30', 'Diagonal 7 # 15-50'],
    'email_domain': ['hotmail.com', 'outlook.com', 'yahoo.com', 'icloud.com', 'aol.com']
}

def generador_preguntas(user_data: dict):
    """
    Genera 5 preguntas de validación basadas en los datos del usuario.
    """
    preguntas = []
    
    # 1. Ciudad de expedición
    ciudad_correcta = user_data.get('ciudad_exp')
    if ciudad_correcta:
        opciones = random.sample([c for c in DUMMY_DATA['ciudad_exp'] if c.lower() != ciudad_correcta.lower()], 3)
        opciones.append(ciudad_correcta)
        random.shuffle(opciones)
        preguntas.append({
            'texto': 'Su cédula fue expedida en:',
            'opciones': opciones,
            'correctas': ciudad_correcta,
            'nombre': 'ciudad_exp'
        })

    # 2. Tipo de crédito
    credito_correcto = user_data.get('tipocredito')
    if credito_correcto:
        opciones = random.sample([c for c in DUMMY_DATA['tipocredito'] if c.lower() != credito_correcto.lower()], 3)
        opciones.append(credito_correcto)
        random.shuffle(opciones)
        preguntas.append({
            'texto': '¿Cuál es su tipo de crédito actual con la entidad?',
            'opciones': opciones,
            'correctas': credito_correcto,
            'nombre': 'tipocredito'
        })

    # 3. Dirección (selección múltiple)
    direccion_correcta = user_data.get('direccion')
    if direccion_correcta:
        opciones = random.sample([d for d in DUMMY_DATA['direccion'] if d.lower() != direccion_correcta.lower()], 3)
        opciones.append(direccion_correcta)  # La respuesta correcta es la dirección completa
        random.shuffle(opciones)
        preguntas.append({
            'texto': '¿Cuál de estas es su dirección registrada?',
            'opciones': opciones,
            'correctas': direccion_correcta,
            'nombre': 'direccion'
        })

    # 4. Número completo de celular
    celular_correcto = user_data.get('celular')
    if celular_correcto and len(str(celular_correcto)) == 10:
        numero_correcto = str(celular_correcto)
        dummy_numbers = {str(random.randint(3000000000, 3599999999)) for _ in range(3)}
        dummy_numbers.add(numero_correcto)
        opciones = list(dummy_numbers)
        random.shuffle(opciones)
        preguntas.append({
            'texto': '¿Cuál es su número de celular registrado completo?',
            'opciones': opciones,
            'correctas': numero_correcto,
            'nombre': 'celular_completo'
        })

    # 5. Email (correo completo)
    email_correcto = user_data.get('mail')
    nombre = user_data.get('nombre', '')
    if email_correcto and '@' in email_correcto:
        usuario_real, dominio_real = email_correcto.split('@')
        variantes = set()

        # Variante 1: parte del nombre + números del usuario real
        if nombre:
            nombre_simple = nombre.split()[0].lower()
            variantes.add(f"{nombre_simple}{usuario_real[-3:]}@{dominio_real}")

        # Variante 2: parte del usuario real
        variantes.add(f"{usuario_real[:-2]}@{dominio_real}")

        # Variante 3: nombre común + dominio real
        # variantes.add(f"carlosortiz@{dominio_real}")

        # Variante 4: nombre + número
        if nombre:
            nombre_simple = nombre.split()[0].lower()
            variantes.add(f"{nombre_simple}258@{dominio_real}")

        # Eliminar la variante igual al correo real
        variantes.discard(email_correcto)
        # Seleccionar solo 3 variantes
        opciones = random.sample(list(variantes), min(3, len(variantes)))
        opciones.append(email_correcto)
        random.shuffle(opciones)
        preguntas.append({
            'texto': '¿Cuál de estos es su correo electrónico registrado?',
            'opciones': opciones,
            'correctas': email_correcto,
            'nombre': 'mail'
        })

    return preguntas

#! --- Vistas ---

def vista_validacion(request):
    if request.method == 'POST':
        cedula = request.POST.get('cedula')
        fecha_expedicion = request.POST.get('fecha_expedicion')
        # fecha_dt = datetime.strptime(fecha_expedicion1, "%Y-%m-%d")
        # fecha_expedicion = fecha_dt.strftime("%d/%m/%Y")
        print(fecha_expedicion)
    
        if not cedula or not fecha_expedicion:
            return render(request, 'validator/initial_form.html', {'error_message': 'Por favor, complete ambos campos.'})

        fecha_expedicion_oracle = fecha_expedicion.replace('-', '/')
        print("fecha formateada ",fecha_expedicion_oracle)
        user_data = get_user_data(cedula, fecha_expedicion_oracle)

        if user_data:
            questions = generador_preguntas(user_data)
            user_info = {
                'nombre': user_data.get('nombre'),
                'cedula': user_data.get('cedula')
            }

            validation_record = ValidationRecord.objects.create(
                cedula=cedula,
                fecha_expedicion=fecha_expedicion,
                user_info=user_info
            )
            
            # Guardar los datos necesarios en la sesión
            request.session['validation_questions'] = questions
            request.session['user_info'] = user_info
            request.session['validation_record_id'] = validation_record.id
            return redirect('titularidad')
        else:
            return render(request, 'validator/initial_form.html', {
                'error_message': 'El asociado no existe o los datos son incorrectos.'
            })

    return render(request, 'validator/initial_form.html')

def vista_preguntas(request):
    questions = request.session.get('validation_questions')
    user_info = request.session.get('user_info')
    validation_record_id = request.session.get('validation_record_id')

    if not questions or not user_info or not validation_record_id:
        return redirect('initial_form')

    if request.method == 'POST':
        correct_answers_count = 0
        user_answers = {}
        print("\n--- INICIO DE VALIDACIÓN DE RESPUESTAS ---")
        for i, q in enumerate(questions):
            user_answer = request.POST.get(f'question_{i}')
            user_answers[q['nombre']] = user_answer

            # --- Bloque de Depuración ---
            print(f"\n[ Pregunta: {q['nombre']} ]")
            print(f"  -> Tu respuesta: '{user_answer}' (Tipo: {type(user_answer)})")
            correct_answer = str(q['correctas'])
            print(f"  -> Respuesta correcta: '{correct_answer}' (Tipo: {type(correct_answer)})")

            cleaned_user_answer = user_answer.lower().strip() if user_answer else None
            cleaned_correct_answer = correct_answer.lower().strip()

            print(f"  -> Comparando: '{cleaned_user_answer}' == '{cleaned_correct_answer}'")
            
            is_correct = cleaned_user_answer is not None and cleaned_user_answer == cleaned_correct_answer
            
            print(f"  -> Resultado: {'CORRECTO' if is_correct else 'INCORRECTO'}")
            # --- Fin Bloque de Depuración ---

            if is_correct:
                correct_answers_count += 1
        
        validation_success = correct_answers_count >= 3
        print(f"\n--- FIN DE VALIDACIÓN ---")
        print(f"Total de respuestas correctas: {correct_answers_count} de {len(questions)}")
        print(f"Resultado final de la validación: {'ÉXITO' if validation_success else 'FALLIDA'}\n")

        try:
            validation_record = ValidationRecord.objects.get(id=validation_record_id)
            validation_record.user_answers = user_answers
            validation_record.validation_success = validation_success
            validation_record.save()
        except ValidationRecord.DoesNotExist:
            pass
        
        request.session.pop('validation_questions', None)
        request.session.pop('user_info', None)
        request.session.pop('validation_record_id', None)

        return render(request, 'validator/result.html', {'validation_success': validation_success})

    return render(request, 'validator/questions_form.html', {
        'preguntas': questions,
        'user_info': user_info
    })

def export_records(request):
    """
    Exportar los datos de ValidationRecord a un archivo XLSX, expandiendo los campos JSON en columnas separadas.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="registros_validacion.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Registros'

    #? --- Creacion Dinámica de Encabezados ---
    
    #? Obtener todos los registros primero para determinar todas las claves JSON posibles 
    records = ValidationRecord.objects.all()
    
    #? Encabezados fijos
    headers = [
        'ID', 'Cédula', 'Fecha de Expedición', 'Validación Exitosa', 'Fecha y Hora'
    ]
    
    #? Encontrar dinamicamente todas las claves de los campos JSON
    info_keys = set()
    answer_keys = set()
    
    for record in records:
        if isinstance(record.user_info, dict):
            info_keys.update(record.user_info.keys())
        if isinstance(record.user_answers, dict):
            answer_keys.update(record.user_answers.keys())

    #? Crear encabezados ordenados y con prefijos para mayor claridad

    sorted_info_keys = sorted(list(info_keys))
    sorted_answer_keys = sorted(list(answer_keys))
    
    headers.extend([f"Info: {key}" for key in sorted_info_keys])
    headers.extend([f"Respuesta: {key}" for key in sorted_answer_keys])
    
    #? Escribir encabezados en la primera fila
    worksheet.append(headers)

    #? --- Escribir los datos de cada registro ---
    
    for record in records:
        row_data = [
            record.id,
            record.cedula,
            record.fecha_expedicion,
            "Sí" if record.validation_success else "No",
            record.timestamp.strftime("%Y-%m-%d %H:%M:%S") if record.timestamp else ""
        ]
        
        #? Añadir datos de user_info, asegurando que el orden coincida con los encabezados
        user_info = record.user_info if isinstance(record.user_info, dict) else {}
        for key in sorted_info_keys:
            row_data.append(user_info.get(key, ''))
            
        #? Añadir datos de user_answers, asegurando que el orden coincida con los encabezados
        user_answers = record.user_answers if isinstance(record.user_answers, dict) else {}
        for key in sorted_answer_keys:
            row_data.append(user_answers.get(key, ''))
            
        worksheet.append(row_data)

    #? Autoajustar el ancho de las columnas para mejor legibilidad
    for i, column_cells in enumerate(worksheet.columns):
        max_length = 0
        column = get_column_letter(i + 1)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response
